print("jamstr4441")

import os
import csv
import openpyxl
from openpyxl.chart import Reference, PieChart
import matplotlib.pyplot as plt
from datetime import date


# Function that asks the user for five numbers and totals them.
def askUser():
    total = 0
    # LOOP LOGIC: This loop runs exactly five times (one pass per number).
    # On each pass it prompts the user, converts their text input to an int,
    # and adds it to the running 'total' accumulator. After five passes the
    # accumulated sum is displayed to the user.
    for i in range(5):
        num = int(input("Enter a number: "))
        total = total + num
    print("The total of those five numbers is:", total)


# Function that asks for five people's names and incomes and appends them to final.csv.
def askIncome():
    # LOOP LOGIC: This loop runs five times, once per person. Each pass collects
    # a name and an annual income, then writes that pair as a new row to the
    # existing final.csv. We open the file in append mode ("a") so we add to the
    # data already there instead of overwriting it. newline="" lets the csv
    # module control line endings so we don't get blank rows between entries.
    with open("C:\\FinalExam\\final.csv", "a", newline="") as f:
        writer = csv.writer(f)
        for i in range(5):
            name = input("Enter a name: ")
            income = input("Enter that person's annual income: ")
            writer.writerow([name, income])


# Function that builds a pie chart in Excel from the final.csv data.
def excelPie():
    # Create a brand-new workbook object in memory to hold the data + chart.
    wb = openpyxl.Workbook()
    # Grab the active worksheet so we have somewhere to write the rows.
    ws = wb.active
    # Open the csv so we can copy its contents into the spreadsheet.
    with open("C:\\FinalExam\\final.csv", "r") as f:
        # Use the csv reader to split each line on commas into [name, income].
        reader = csv.reader(f)
        # Walk every row in the file one at a time.
        for row in reader:
            # Skip any blank/empty lines so they don't create junk rows.
            if len(row) < 2:
                continue
            # Write the name in col A and the income CAST TO INT in col B.
            # Casting is required or the pie chart plots nothing (blank).
            ws.append([row[0], int(row[1])])
    # Build a Reference to the numeric data (column B = col 2) for the chart.
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    # Build a Reference to the category labels (column A = the names).
    labels = Reference(ws, min_col=1, min_row=1, max_row=ws.max_row)
    # Create the PieChart object that we'll attach the data to.
    pie = PieChart()
    # Feed the income values into the chart as the data series to plot.
    pie.add_data(data)
    # Attach the names as the slice labels so each wedge is identified.
    pie.set_categories(labels)
    # Set the chart title to StudentID + today's date, per the instructions.
    pie.title = "jamstr4441 " + date.today().strftime("%B %d, %Y")
    # Drop the chart's top-left corner at cell D1 so it sits beside the data.
    ws.add_chart(pie, "D1")
    # Save everything (data + pie chart) into the new file final.xlsx.
    wb.save("C:\\FinalExam\\final.xlsx")


# Function that builds a vertical bar graph from the same data using matplotlib.
def verticalBar():
    names = []
    incomes = []
    # Read the same csv and split it into parallel lists for plotting.
    with open("C:\\FinalExam\\final.csv", "r") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 2:
                continue
            names.append(row[0])
            incomes.append(int(row[1]))  # cast to int so bar heights are numeric
    # plt.bar() draws a VERTICAL bar chart: names on x-axis, incomes as heights.
    plt.bar(names, incomes)
    # Title is StudentID + today's date, matching the pie chart requirement.
    plt.title("jamstr4441 " + date.today().strftime("%B %d, %Y"))
    plt.xlabel("Name")
    plt.ylabel("Annual Income")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()


# Run the program
askUser()
askIncome()
excelPie()
verticalBar()
